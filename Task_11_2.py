import openpyxl
from openpyxl import Workbook
wb = Workbook()
wb.save('excel3.xlsx')
wb = openpyxl.load_workbook('excel3.xlsx')
wb.create_sheet('Лист1')
print(wb.sheetnames)
ws = wb['Лист1']

import csv
with open('test3.csv', encoding='utf-8') as file:
    rows = csv.reader(file)
    su = 0
    lst = []
    for row in rows:
        lst.append(row)
    for i in lst[1:]:
        su += int(i[4])
    lst0 = lst[0]
    newlst=sorted(lst[1:], key=lambda x: (x[3],x[1],x[2]))
    print(lst0, newlst, su)

ws.append(lst0)
for k in newlst:
    ws.append(k)
ws.cell(ws.max_row + 1, 1).value = "Итого"
ws.cell(ws.max_row, 5).value = su
wb.save('excel3.xlsx')

