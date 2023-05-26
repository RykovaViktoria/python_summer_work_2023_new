# dct = {1:{11:{111:1111}},2:{22:{222:2222}},3:{33:333},4:444}
# for k,v in dct.items():
#     print(k,v)
#     if type(v) == dict:
#         for p,q in v.items():
#             print(p, q)
#             if type(q) == dict:
#                 for m,n in q.items():
#                     print(m,n)
#
# не использовать букву l, т.к. похожа на 1.

# Классы пишутся с большой буквы.

# from functools import cmp_to_key
# def xy(x, y):
#     if x < y: return -1
#     elif x==y: return 0
#     else: return 1
#
# lst = [100, -1, 0, 1000, 20, 5]
# print(sorted(lst, key = cmp_to_key(xy)))

# x = 1235
# i = 2
# while x % i: #если НЕ делится
#     i += 1
# print(i)
#
# from functools import cmp_to_key
# def pp(x, y):
#     print(x,y)
#     i = 2
#     while x % i: #если НЕ делится
#         i += 1
#     j = 2
#     while y % j: #если НЕ делится
#         j += 1
#     if i < j: return -1
#     elif i==j: return 0
#     else: return 1
# lst = [31,5,7,2,9,10,11,23,125,1024]
# print(sorted(lst, key = cmp_to_key(pp)))

# from functools import reduce
# print(reduce(lambda x,y:x*y,[1,2,3,4,5]))
#
# import functools
# print(dir(functools))
#
# partial


# Работа с файлами:
# Удобно читать построчно:
# # for line in f:
# #     print(line)
#
# print(file=f)

# f = open("text.txt", 'w', encoding = 'utf-8')
# for i in range(10):
#     print(i, i*i, file = f)
#     print(111, i, i * i) - посмотерть, что записать
# f.close()
# f = open("text.txt", 'r', encoding = 'utf-8')
# for i in f:
#     print(222, i.strip())
# print(f)
#
# # Менеджер контекста
# # Чтобы не писать f.close():
# # a - дозапись
#
# with open("text.txt", 'a', encoding = 'utf-8') as fout:
#     for i in range(10):
#         print(i, i*i, file = fout)
#         print(111, i, i * i)
#
# with open("text.txt", 'r', encoding = 'utf-8') as fin:
#     for i in fin:
#         print(222, i.strip())


# with open("text.txt", 'w', encoding = 'utf-8') as fout:
#     for i in range(5):
#         s = input()
#         print(s, file = fout)
# Рабоче-крестьянский вариант:

# with open("text.txt", 'r', encoding = 'utf-8') as fin:
#     s = fin.readlines()
#
# with open("text2.txt", 'w', encoding = 'utf-8') as fout:
#     for i in s:
#         x = ''.join(sorted(i.strip().split(), key = str.lower))
#         print(x, file = fout)
# with open("text2.txt", encoding='utf-8') as fi:
#     for i in fi:
#         print(i.strip())

# Модуль openpyxl для работы с Excel

# import openpyxl
# from openpyxl import Workbook
# wb = Workbook()
# wb.save('test_excel.xlsx')

# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb.active
# print(ws.title)
# wb.create_sheet('New')
# print(wb.sheetnames)
# wb.save('test_excel.xlsx')

# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb.active
# ws['A1'].value = 123
# print(ws['A1'].value)
# ws['B2'].value=234
# c=ws['B2']
# print(c.row)
# print(c.column)
# print(c.coordinate)

# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb.active
# ws['A1'].value = 100
# ws['B2'].value=200
# ws['C3'].value = ws['A1'].value + ws['B2'].value
# ws.cell(3,3).value
# for i in range(1,5):
#     print(i, ws.cell(i,2).value)
# for cellObj in ws['A1':'C3']:
#     for cell in cellObj:
#         print(cell.coordinate, cell.value)

# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb.active
# ws['A1'].value = 100
# ws['B2'].value=200
# ws['C3'].value = ws['A1'].value + ws['B2'].value
# ws.cell(3,3).value
# for i in range(1, ws.max_row+1):
#     for j in range(1, ws.max_column+1):
#             print(i,j, ws.cell(i,j).value)
#
# Прочитать всю страницу. Преобразовать. и потом перезаписать на новую страницу.

# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb.active
# ws['A1'].value = 100
# ws['B2'].value=200
# ws['C3'].value = ws['A1'].value + ws['B2'].value
# ws.cell(3,3).value
# for i in range(1, ws.max_row+1):
#     for j in range(1, ws.max_column+1):
#             print(i,j, ws.cell(i,j).value)
#
# Прочитать всю страницу. Преобразовать. и потом перезаписать на новую страницу.

# import openpyxl
# from openpyxl import Workbook
# wb = openpyxl.load_workbook('test_excel.xlsx')
# s=wb.sheetnames
# res=[]
# for i in s:
#     ws = wb[i]
#     res.append((ws.title, ws.max_row, *ws.max_column))
# print(sorted(res))
# print(sorted(res, key = lambda x: -x[1]))

# import openpyxl
# from openpyxl import Workbook
# wb = openpyxl.load_workbook('test_excel.xlsx')
# s=wb.sheetnames
# print(s)
# res=[]
# for i in s:
#     ws = wb[i]
#     res.append((ws.title, ws.max_row, *ws.max_column))
# print(sorted(res))
# print(sorted(res, key = lambda x: -x[1]))


# import openpyxl
# wb = openpyxl.load_workbook('test_excel.xlsx')
# print(wb.sheetnames)
# ws = wb['New']
# ws['A1'].value = 100
# ws['B2'].value=200
# ws['C3'].value = ws['A1'].value + ws['B2'].value
# su = 0
# for i in range(1, ws.max_row+1):
#     su += ws.cell(i, 2).value
# j=ws.max_row+1
# ws.cell(j,1).value='ИТОГО'
# ws.cell(j,2).value=su
# print(su)
# wb.save('test_excel.xlsx')

import openpyxl

wb = openpyxl.load_workbook("text_excel.xlsx")
ws = wb["Лист1"]
print((ws.title, ws.max_row * ws.max_column))
su = 0
for i in range(1, ws.max_row + 1):
    su += ws.cell(i, 2).value
j = ws.max_row + 1
print(su)
ws.cell(ws.max_row + 1, 1). value = "Итого"
ws.cell(j, 2).value = su
wb.save("text.xlsx")


