import openpyxl
wb = openpyxl.load_workbook('excel.xlsx')
# wb.create_sheet('Итого')
print(wb.sheetnames)
ws = wb.active
wn = wb['Итого']
workers = []
for i in range(1, ws.max_row+1):
    for j in range(1, ws.max_column+1):
        workers.append(ws.cell(i,j).value)
print(workers)
d = {}
i = 0
while len(workers) > i:
    if workers[i] in d:
        d[workers[i]] += workers[i+1]
    else: d[workers[i]] = workers[i+1]
    i+=2
print(d)
su = 0
for k,v in d.items():
    su += v
print(su)
j=1
for k,v in d.items():
    wn.cell(j, 1).value=k
    wn.cell(j, 2).value=v
    j+=1
wn.cell(wn.max_row + 1, 1).value = "Итого"
wn.cell(wn.max_row, 2).value = su
for i in range(1, wn.max_row+1):
    for j in range(1, wn.max_column+1):
            print(i,j, wn.cell(i,j).value)

wb.save('excel.xlsx')